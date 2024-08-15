import sys

from Receiver.can_dbc_decoder import DbcDecoder

if __name__ == "__main__":  # Warn if trying to run this as a script
    print("\n**********************************************")
    print("   This is not meant to be run as a main script")
    print("   Run log-to-data.py or live-telem.py instead")
    print("**********************************************\n")
    sys.exit(4)

from datetime import datetime, timedelta, timezone
from openpyxl import load_workbook
from os.path import exists as fileExists
import struct
from crccheck.crc import Crc16Modbus
from binascii import hexlify
from functools import lru_cache, cache

import numpy as np
from numpy import uint32
import logging
from Receiver.receiver_config import configFile

logger = logging.getLogger(__name__)
logger.addHandler(logging.NullHandler())


class TelemetryParser:
    def __init__(self, config_file=configFile):
        self.lastGPSTime: datetime = datetime.now(
            timezone.utc
        )  # use this by default until the pi rtc goes out of sync (i.e the UPS fails). It can be resynced if reconnected to internet.
        self.timeFetched: uint32 = uint32(
            0
        )  # Time since time variables were last updated in seconds #round(time.time() * 1000). Using numpy to force unsigned and integer overflows are needed
        self.rowForCurrentMessage = dict()
        self.config_file = config_file
        self.decoder = DbcDecoder()

    @property
    @cache
    def config(self):
        config = dict()

        # CONFIG INIT REGION
        # get config from Excel config workbook. Check 'CAN data' sheet exists

        config_file = self.config_file
        if not fileExists(config_file):
            logger.info("Config file " + config_file + " does not exist")
            sys.exit(5)

        logger.debug("Loading config (this process takes a while)")
        configBook = load_workbook(
            config_file, read_only=True, keep_vba=True, data_only=True
        )  # Config has VBA macros. Also, data_only=True only works if formulae were evaluated. This is fine since excel does this and openpyxl does not modify config
        if not "CAN Data" in configBook.sheetnames:  # check CAN data worksheet exists
            logger.warning(
                "Missing 'CAN Data' worksheet in workbook. Is the config file correct?"
            )
            logger.warning(
                "Worksheets available: " + str(configBook.sheetnames))
            sys.exit(1)
        configSheet = configBook["CAN Data"]
        configColumns = {cell.value: i + 1 for i, cell in
                         enumerate(configSheet[1])}
        for row in configSheet.iter_rows(min_row=2):
            if row[0].value == "END":
                break
            rowAsDict = {
                columnName: row[configColumns[columnName] - 1].value
                for columnName in configColumns
            }
            config[row[configColumns["CAN_ID (dec)"] - 1].value] = rowAsDict
        configBook.close()
        # print config
        for address, packet_config in config.items():
            logger.info("{}\t{}".format(hex(address), packet_config))
        return config

    def __get_time(self, recievedMillis: uint32) -> datetime:
        millisDelta: uint32 = recievedMillis - self.timeFetched
        if recievedMillis < self.timeFetched:
            millisDelta = (
                    millisDelta + 2 ** 32
            )  # Unsign the delta. This method should work as long as the GPS update is not older than 2^32-1 milliseconds

        currentTime = self.lastGPSTime + timedelta(
            milliseconds=millisDelta.item())
        logger.debug(
            "millisDelta: " + str(millisDelta.item()) + " -> "
                                                        "Current Time: " + currentTime.strftime(
                "%Y-%m-%d %H:%M:%S.%f")
        )
        return currentTime

    def __from_config(self, dictKey):
        if dictKey in self.rowForCurrentMessage:
            return self.rowForCurrentMessage[dictKey]
        else:
            logger.info(
                "Column '" + dictKey + "' missing in 'CAN Data' worksheet in config"
            )
            sys.exit(2)

    # TRANSLATE MESSAGE REGION
    def translate_msg(
            self, msgBytesAndTime: bytearray
    ) -> tuple[
        str, str, dict, datetime, bool
    ]:  # Format: TI0 TI1 TI2 TI3 ID0 ID1 DLC B0 B1 B2 B3 B4 B5 B6 B7 CRC0 CRC1 (NOTE that end of frame marker is not included)
        logger.debug("Translating -> " + str(msgBytesAndTime))

        msgBytes = msgBytesAndTime[4:]

        # CRC check
        msgCRCStatus = self.__check_crc(msgBytesAndTime)
        if not msgCRCStatus:
            logger.debug(
                "CRC FAILED (ignoring message) "
            )  # + msgTime.strftime("%Y-%m-%d %H:%M:%S"))
            return (
                "CRCFail",
                "",
                {"Data": hexlify(msgBytesAndTime)},
                datetime(1970, 1, 1, 3, 0, 0),
                False,
            )

        # convert recieved millis delta time
        recievedMillisTime = np.frombuffer(
            msgBytesAndTime[0:4], dtype=uint32
        )  # int.from_bytes(msgBytesAndTime[0:3], byteorder="little")
        msgTime = self.__get_time(recievedMillisTime)

        # do a lookup in spreadsheet using can id to work out can message type
        canId = msgBytes[0] << 8 | msgBytes[1]
        if canId in self.config.keys():
            self.rowForCurrentMessage = self.config[canId]
        else:
            logger.exception(
                "Error. Could not config entry for id " + str(canId))
            return "ID UNRECOGNISED", "ERROR", {
                "ID": canId}, msgTime, msgCRCStatus

        # Translate
        msgItem: str = self.__from_config("ItemCC")
        msgSource: str = self.__from_config("SourceCC")
        msgBody = self.decoder.decode_can_msg(canId, msgBytes[3:])
        # if canId == 0x0F6 and msgBody["GpsDay"] != 0:
        if canId == 24 and msgBody["GpsDay"] != 0:
            self.update_last_gps_time(msgBody,recievedMillisTime)
        return msgItem, msgSource, msgBody, msgTime, msgCRCStatus

    def update_last_gps_time(self, msgBody, recievedMillisTime):
        logger.debug("Updating GPS time...")
        try:
            self.lastGPSTime = datetime(
                hour=msgBody["GpsHour"],
                minute=msgBody["GpsMinute"],
                second=msgBody["GpsSeconds"],
                day=msgBody["GpsDay"],
                month=msgBody["GpsMonth"],
                year=2000 + msgBody["GpsYear"],
                tzinfo=timezone.utc,
            )  # msgData only contains last 2 digits of year so have to add 2000
            self.timeFetched = (
                recievedMillisTime  # update when data was last fetched
            )
            logger.debug(
                "GPS time is now: " + self.lastGPSTime.strftime(
                    "%Y-%m-%d %H:%M:%S")
            )
        except ValueError:
            logger.exception("Invalid values for GPS time, %s",
                             str(msgBody))

    @staticmethod
    def __check_crc(msgBytes: bytearray) -> bool:
        data = msgBytes[0:-2]  # All except the last two
        crc_rcvd = int.from_bytes(
            msgBytes[-2:], "big"
        )  # Convert received CRC code to int
        crc_calc = Crc16Modbus.calc(data)
        return crc_calc == crc_rcvd
