import sys

if __name__ == '__main__':  # Warn if trying to run this as a script
    print("\n**********************************************")
    print("   This is not meant to be run as a main script")
    print("   Run log-to-data.py or live-telem.py instead")
    print("**********************************************\n")
    sys.exit(4)

from datetime import datetime, timedelta, timezone
from openpyxl import load_workbook
from os.path import exists as fileExists
import struct
from crccheck.crc import Crc16Modbus
from binascii import hexlify
# from csv import DictReader as csvAsDictReader
import numpy as np
from numpy import uint32
import logging

logger = logging.getLogger(__name__)
logger.addHandler(logging.NullHandler())

from Receiver.receiver_config import configFile

# TIME REGION
# lastGPSTime: datetime = datetime(year=1970, month=1, day=1, hour=3, minute=0, second=0, tzinfo=timezone.utc) #Excel does not support timezones tzinfo=timezone.utc
lastGPSTime: datetime = datetime.now(
    timezone.utc)  # use this by default until the pi rtc goes out of sync (i.e the UPS fails). It can be resynced if reconnected to internet.
# If out of sync AND the GPS rtc has failed, use line above and manually try and set a time close to the current time.

timeFetched: uint32 = uint32(
    0)  # Time since time variables were last updated in seconds #round(time.time() * 1000). Using numpy to force unsigned and integer overflows are needed


def __getTime(recievedMillis: uint32) -> datetime:
    millisDelta: uint32 = recievedMillis - timeFetched
    if recievedMillis < timeFetched:
        millisDelta = millisDelta + 2 ** 32  # Unsign the delta. This method should work as long as the GPS update is not older than 2^32-1 milliseconds

    currentTime = lastGPSTime + timedelta(milliseconds=millisDelta.item())
    logger.debug("millisDelta: " + str(millisDelta.item()) + " -> "
                                                             "Current Time: " + currentTime.strftime(
        "%Y-%m-%d %H:%M:%S.%f"))
    return currentTime

#CONFIG INIT REGION
#get config from excel config workbook. Check 'CAN data' sheet exists
if not fileExists(configFile):
    logger.info("Config file " + configFile + " does not exist")
    sys.exit(5)
# configData = open(configFile)

# with open(configFile) as configData:
logger.debug("Loading config (this process takes a while)")
config: dict[int, dict] = dict()
# with load_workbook(configFile, read_only=True, keep_vba=True, data_only=True) as configBook:
configBook = load_workbook(configFile, read_only=True, keep_vba=True,
                           data_only=True)  # Config has VBA macros. Also, data_only=True only works if formulae were evaluated. This is fine since excel does this and openpyxl does not modify config
if not 'CAN Data' in configBook.sheetnames:  # check CAN data worksheet exists
    logger.warning(
        "Missing 'CAN Data' worksheet in workbook. Is the config file correct?")
    logger.warning("Worksheets available: " + str(configBook.sheetnames))
    sys.exit(1)
configSheet = configBook['CAN Data']
# c = csvAsDictReader(configData)
configColumns = {cell.value: i + 1 for i, cell in enumerate(configSheet[1])}
can_ID_col = configColumns["CAN_ID (dec)"]
for row in configSheet.iter_rows(min_row=2):
    if row[0].value == "END":
        break
    rowAsDict = {columnName: row[configColumns[columnName] - 1].value for
                 columnName in configColumns}
    config[row[configColumns["CAN_ID (dec)"] - 1].value] = rowAsDict
configBook.close()

rowForCurrentMessage = dict()
def __fromConfig(dictKey):
    # row = config.get(canId)
    if dictKey in rowForCurrentMessage:
        return rowForCurrentMessage[dictKey]
    else:
        logger.info(
            "Column '" + dictKey + "' missing in 'CAN Data' worksheet in config")
        sys.exit(2)


# print config
for address, packet_config in config.items():
    logger.info("{}\t{}".format(hex(address), packet_config))


# TRANSLATE MESSAGE REGION
def translateMsg(msgBytesAndTime: bytearray) -> tuple[
    str, str, dict, datetime, bool]:  # Format: TI0 TI1 TI2 TI3 ID0 ID1 DLC B0 B1 B2 B3 B4 B5 B6 B7 CRC0 CRC1 (NOTE that end of frame marker is not included)
    logger.debug("Translating -> " + str(msgBytesAndTime))

    msgBytes = msgBytesAndTime[4:]

    # CRC check
    msgCRCStatus = __checkCRC(msgBytesAndTime)
    if not msgCRCStatus:
        logger.debug(
            "CRC FAILED (ignoring message) ")  # + msgTime.strftime("%Y-%m-%d %H:%M:%S"))
        return "CRCFail", "", {"Data": hexlify(msgBytesAndTime)}, datetime(1970,
                                                                           1, 1,
                                                                           3, 0,
                                                                           0), False

    # convert recieved millis delta time
    recievedMillisTime = np.frombuffer(msgBytesAndTime[0:4],
                                       dtype=uint32)  # int.from_bytes(msgBytesAndTime[0:3], byteorder="little")
    msgTime = __getTime(recievedMillisTime)

    # do a lookup in spreadsheet using can id to work out can message type
    canId = msgBytes[0] << 8 | msgBytes[1]
    global rowForCurrentMessage
    try:
        rowForCurrentMessage = config[canId]
    except KeyError:
        logger.exception("Error. Could not config entry for id " + str(canId))
        return "ID UNRECOGNISED", "ERROR", {"ID": canId}, msgTime, msgCRCStatus

    # Translate
    msgItem: str = __fromConfig("ItemCC")
    msgSource: str = __fromConfig("SourceCC")
    msgDLC = __fromConfig("DLC")
    msgData = struct.unpack(__fromConfig("struct unpack code"),
                            msgBytes[3:(3 + msgDLC)])
    msgBody: dict = {}
    dataIterator = 0
    for i in range(msgDLC):
        fieldValue = __fromConfig("BYTE_" + str(i) + "CC")
        if not (
                fieldValue == '-'):  # if byte labelled '-', then is part of a value that spans multiple bytes or is not being used at all
            msgBody.update({fieldValue: msgData[dataIterator]})
            dataIterator = dataIterator + 1

    logger.debug(
        f'{msgSource}: {msgBody}')  # translated data without any extra decoding

    # if GPS time and fix message, update time
    if canId == 24 and msgData[4] != 0:  # can id for GPS Time and Fix message (hex: 0x0F6)
        logger.debug("Updating GPS time...")
        global lastGPSTime
        global timeFetched
        try:
            lastGPSTime = datetime(
                hour=msgData[0],
                minute=msgData[1],
                second=msgData[2],
                day=msgData[3],
                month=msgData[4],
                year=2000 + msgData[5],
                tzinfo=timezone.utc)  # msgData only contains last 2 digits of year so have to add 2000
            timeFetched = recievedMillisTime  # update when data was last fetched
            logger.debug(
                "GPS time is now: " + lastGPSTime.strftime("%Y-%m-%d %H:%M:%S"))
        except ValueError:
            logger.exception("Invalid values for GPS time, %s", str(msgData))

    # mppt
    if canId == 1905 or canId == 1906:
        logger.info("Decoding MPPT")
        newMsgBody: dict = {
            'VoltageIn': ((msgBody["FlagsAndMsbVoltageIn"] & 3) << 8) | msgBody[
                "LsbVoltageIn"],
            # bitwise and with 3 because cannot confirm if other bits (marked 'x') in byte are 0
            'CurrentIn': ((msgBody["MsbCurrentIn"] & 3) << 8) | msgBody[
                "LsbCurrentIn"],
            'VoltageOut': ((msgBody["MsbVoltageOut"] & 3) << 8) | msgBody[
                "LsbVoltageOut"],
            'AmbientTemperature': msgBody["AmbientTemperature"],
            'Flag/BatteryVoltageLevelReached': (
                    (msgBody["FlagsAndMsbVoltageIn"] & 128) >> 7),
            'Flag/OverTemperature': (
                    (msgBody["FlagsAndMsbVoltageIn"] & 64) >> 6),
            'Flag/NoCharge': ((msgBody["FlagsAndMsbVoltageIn"] & 32) >> 5),
            'Flag/UnderVoltage': ((msgBody["FlagsAndMsbVoltageIn"] & 16) >> 4)
        }
        msgBody = newMsgBody
        logger.info("MPPT Decode: " + str(msgBody))

    return msgItem, msgSource, msgBody, msgTime, msgCRCStatus


def __checkCRC(msgBytes: bytearray) -> bool:
    data = msgBytes[0:-2]  # All except the last two
    crc_rcvd = int.from_bytes(msgBytes[-2:],
                              "big")  # Convert received CRC code to int
    crc_calc = Crc16Modbus.calc(data)
    return crc_calc == crc_rcvd
