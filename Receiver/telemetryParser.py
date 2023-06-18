import sys

if __name__ == '__main__':  # Warn if trying to run this as a script
    print("\n**********************************************")
    print("   This is not meant to be run as a main script")
    print("   Run log-to-data.py or live-telem.py instead")
    print("**********************************************\n")
    sys.exit(4)

from datetime import datetime, UTC, timedelta
from openpyxl import load_workbook
from os.path import exists as fileExists
import struct
import time
from crccheck.crc import Crc16Modbus
from binascii import hexlify

#configFile: str = './CANConfig.xslx' #raspberrypi
configFile: str = '../../CANTranslator/config/CANBusData(saved201022)Modified.xlsm' #testing with windows

#TIME REGION
lastGPSTime: datetime = datetime(1970, 1, 1, 3, 0, 0, UTC) #Excel does not support timezones tzinfo=timezone.utc
timeFetched = time.time() #Time since time variables were last updated in seconds #round(time.time() * 1000)
def __getTime() -> datetime:
    #print(lastGPSTime.timestamp() + (time.time() - timeFetched))
    currentTime = lastGPSTime + timedelta(seconds = (time.time() - timeFetched))
    #currentTime = datetime.fromtimestamp(lastGPSTime.timestamp() + (time.time() - timeFetched))
    return currentTime

#CONFIG INIT REGION
#get config from excel config workbook. Check 'CAN data' sheet exists
if not fileExists(configFile):
    print("Config file " + configFile + " does not exist")
    sys.exit(5)
config = load_workbook(configFile, read_only=True, keep_vba=True, data_only=True) #Config has VBA macros. Also, data_only=True only works if formulae were evaluated. This is fine since excel does this and openpyxl does not modify config
if not 'CAN Data' in config.sheetnames: #check CAN data worksheet exists
    print("Missing 'CAN Data' worksheet in workbook. Is the config file correct?")
    print("Worksheets available: ", end='')
    print(config.sheetnames)
    sys.exit(1)
configSheet = config['CAN Data']

#map column labels to column indices
configColumns = dict()
print("Getting column labels")
print([cell.value for cell in configSheet[1][0:configSheet.max_column]])
for columnIterator in range(1, configSheet.max_column + 1): #cell() is 1 indexed (i.e. cell A1 is fetched with (1,1) coords)
    if configSheet.cell(row=1, column=columnIterator).value is None: #If type is 'NoneType', ignore column
        continue
    #print("Column: " + configSheet.cell(row=1, column=columnIterator).value)
    if configSheet.cell(row=1, column=columnIterator).value in configColumns: #check for duplicates
        print("Column labels must be unique in 'CAN Data' worksheet in config")
        sys.exit(3)
    configColumns.update({configSheet.cell(row=1, column=columnIterator).value: columnIterator})
    columnIterator = columnIterator + 1

#TRANSLATE MESSAGE REGION
def translateMsg(msgBytes: bytearray) -> tuple[str, str, dict, datetime, bool]: #Format: ID0 ID1 DLC B0 B1 B2 B3 B4 B5 B6 B7 CRC0 CRC1 (NOTE that end of frame marker is not included)
    print("Translating -> " + str(msgBytes))

    #get time
    msgTime = __getTime() #get current time (according to GPS time, not system time)

    #CRC check
    msgCRCStatus = __checkCRC(msgBytes)
    if not msgCRCStatus:
        print("CRC FAILED at " + msgTime.strftime("%Y-%m-%d %H:%M:%S"))
        return "CRCFail", "", {"Data": hexlify(msgBytes)}, msgTime, False

    #do a lookup in spreadsheet using can id to work out can message type
    canId = msgBytes[0] << 8 | msgBytes[1] #int(''.join(hexNumber for hexNumber in msgBytes[0:2]), base=16)
    configRow = 1
    # for r in config.iter_rows():
    #     if r[__getConfigColumn("CAN_ID (dec)")] == canId:
    #         configRow = r
    #         break
    #     if r[0] == "END":
    #         print("Could not find config row for id: " + str(canId))
    #         return "Unknown CanID", "", {"Data": hexlify(msgBytes)}, msgTime, msgCRCStatus #If this runs, the config being used here is out of date
        
    while not (str(configSheet.cell(configRow, 1).value) == "END"):
        #print(configSheet.cell(configRow, __getConfigColumn("CAN_ID (dec)")).value)
        if configSheet.cell(configRow, __getConfigColumn("CAN_ID (dec)")).value == canId:
            configFound = True
            break
        configRow += 1
    if not configFound:
        print("Could not find config row for id: " + str(canId))
        return "Unknown CanID", "", {"Data": hexlify(msgBytes)}, msgTime, msgCRCStatus #If this runs, the config being used here is out of date
    #print(canId)
    print("Config Row: ", end='')
    print([cell.value for cell in configSheet[configRow][0:configSheet.max_column]])
    print('\n\n')
    
    #Translate
    msgItem: str = configSheet.cell(configRow, __getConfigColumn("ItemCC")).value #use camel case formats to avoid issues with storing data
    msgSource: str = configSheet.cell(configRow, __getConfigColumn("SourceCC")).value
    #msgData = struct.unpack(configSheet.cell(configRow, __getConfigColumn("struct unpack code")).value, bytes.fromhex(''.join(str(element) for element in msgBytes[3:12])))
    msgDLC = configSheet.cell(configRow, __getConfigColumn("DLC")).value
    msgData = struct.unpack(configSheet.cell(configRow, __getConfigColumn("struct unpack code")).value, msgBytes[3:(3 + msgDLC)])
    msgBody: dict = {}
    dataIterator = 0
    for i in range(0, 8):
        fieldName = configSheet.cell(configRow, __getConfigColumn("BYTE_" + str(i) + "CC")).value
        if not (fieldName == '-'): #if byte labelled '-', then is part of a value that spans multiple bytes or is not being used at all
            msgBody.update({fieldName: msgData[dataIterator]})
            dataIterator = dataIterator + 1
    #msgDataType = configSheet.cell(configRow, __getConfigColumn("data type"))

    #if GPS time and fix message, update time
    if canId == 246: #can id for GPS Time and Fix message (hex: 0x0F6)
        print("Updating GPS time...")
        global lastGPSTime
        global timeFetched
        lastGPSTime = datetime( \
            hour = msgData[0], \
            minute = msgData[1], \
            second = msgData[2], \
            day = msgData[3], \
            month = msgData[4], \
            year = 2000 + msgData[5], \
            tzinfo=UTC ) #msgData only contains last 2 digits of year so have to add 2000
        timeFetched = time.time() # update when data was last fetched
        print("GPS time is now: " + lastGPSTime.strftime("%Y-%m-%d %H:%M:%S"))

    return msgItem, msgSource, msgBody, msgTime, msgCRCStatus
    
def __getConfigColumn(columnLabel: str) -> int: #used to make sure reordering of columns in config spreadsheet does not cause errors in this script
    if columnLabel in configColumns:
        return configColumns[columnLabel]
    else:
        print("Column '" + columnLabel + "' missing in 'CAN Data' worksheet in config")
        sys.exit(2)

def __checkCRC(msgBytes: bytearray) -> bool:
    data = msgBytes[0:-2]      # All except the last two
    crc_rcvd = int.from_bytes(msgBytes[-2:], "big")  # Convert received CRC code to int
    crc_calc = Crc16Modbus.calc(data)
    return crc_calc == crc_rcvd
