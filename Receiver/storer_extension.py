import logging
import os

from abc import ABC, abstractmethod
from datetime import datetime
from typing import Union

from openpyxl import load_workbook, Workbook
from os.path import exists as fileExists

from influxdb import InfluxDBClient

logger = logging.getLogger(__name__)
logger.addHandler(logging.NullHandler())


class StorerExtension(ABC):
    """Represents a storage option.

    Each class represents a storage option (e.g. csv, excel, influx...)
    The methods `store_data` and `close` must be implemented.

    Additional methods can be implemented to be called by either `store_data`
    or `close`.
    """

    @abstractmethod
    def store_data(
        self,
        msg_item: str,
        msg_source: str,
        msg_body: dict,
        msg_time: datetime,
        msg_crc_status: bool,
    ) -> None:
        raise NotImplementedError

    @abstractmethod
    def close(self) -> None:
        raise NotImplementedError


class ExcelStorer(StorerExtension):
    """Storer for an Excel file.

    A new Excel will be created if not already present and a sheet named
    `Translated Messages` will be created.

    Each row will correspond to a single CAN message.

    When appending to an existing file, a gap of three lines will be added
    before the new session.
    """

    def __init__(self, xlsx_output_file: Union[str, os.PathLike]):
        self.xlsx_output_file = xlsx_output_file
        # skip first row as that is for column labels
        self.xlsx_out_row_pointer: int = 2
        # if '', disable xlsx output
        if self.xlsx_output_file == "":
            logger.error("No excel output path provided.")
            return

        if fileExists(self.xlsx_output_file):
            self.xlsx_out_workbook = load_workbook(
                self.xlsx_output_file, read_only=False
            )
            if "Translated Messages" in self.xlsx_out_workbook.sheetnames:
                self.xlsx_out_worksheet = self.xlsx_out_workbook["Translated Messages"]
                self.xlsx_out_row_pointer = self.xlsx_out_worksheet.max_row + 3
                # leave gap in between sessions. Just to make reader aware
                # that more than one session continued with the same file
            else:
                self.create_new_sheet()
        else:
            self.xlsx_out_workbook = Workbook()  # Create new workbook
            self.create_new_sheet()

    def create_new_sheet(self):
        # Create the new sheet
        self.xlsx_out_worksheet = self.xlsx_out_workbook.create_sheet(
            title="Translated Messages"
        )

        # Add columns labels at the top of the sheet
        for i, label in enumerate(["Date", "Time", "Source", "Item", "Data..."]):
            self.xlsx_out_worksheet.cell(column=i + 1, row=1, value=label)

        # longest record is 19 cells so put CRC in 20th
        self.xlsx_out_worksheet.cell(column=21, row=1, value="CRC check")

        # Add filters to columns (Source and Item)
        self.xlsx_out_worksheet.auto_filter.ref = "A1:T" + str(
            self.xlsx_out_worksheet.max_row
        )

        # Set the pointer to write at the first row under the labels
        self.xlsx_out_row_pointer = 2

    def store_data(
        self,
        msg_item: str,
        msg_source: str,
        msg_body: dict,
        msg_time: datetime,
        msg_crc_status: bool,
    ) -> None:
        xlsx_out_workbook = self.xlsx_out_workbook
        xlsx_out_worksheet = self.xlsx_out_worksheet
        xlsx_out_row_pointer = self.xlsx_out_row_pointer
        xlsx_output_file = self.xlsx_output_file

        msg_time = msg_time.replace(tzinfo=None)
        xlsx_out_worksheet.cell(column=1, row=xlsx_out_row_pointer, value=msg_time)
        xlsx_out_worksheet.cell(column=2, row=xlsx_out_row_pointer, value=msg_time)

        xlsx_out_worksheet.cell(column=3, row=xlsx_out_row_pointer, value=msg_source)
        xlsx_out_worksheet.cell(column=4, row=xlsx_out_row_pointer, value=msg_item)

        column_pointer: int = 5
        for data_label, value in msg_body.items():
            xlsx_out_worksheet.cell(
                column=column_pointer, row=xlsx_out_row_pointer, value=data_label
            )
            xlsx_out_worksheet.cell(
                column=column_pointer + 1, row=xlsx_out_row_pointer, value=value
            )
            column_pointer += 2

        xlsx_out_worksheet.cell(
            column=21, row=xlsx_out_row_pointer, value=msg_crc_status
        )

        # save every time a line is written. NOTE: this can corrupt xlsx file if
        # program doesn't exit gracefully (i.e. the computer is unplugged).
        # This is fine as it can be rebuilt using telemetry hex dump on SD card
        self.xlsx_out_row_pointer += 1
        xlsx_out_workbook.save(xlsx_output_file)

    def close(self) -> None:
        logger.info("Closing xlsx output file")
        self.xlsx_out_workbook.save(self.xlsx_output_file)
        self.xlsx_out_workbook.close()


class Influx1Storer(StorerExtension):
    """Storer for Influx db 1.x

    Storer for Influx db 1.x utilising the python-influxdb library.
    This storer does not support Influx 2 or 3.
    """

    def __init__(self, if_credentials):
        # if false, do not initialise or enable influx output
        if if_credentials.enabled:
            self.influx_client = InfluxDBClient(
                host=if_credentials.host,
                port=if_credentials.port,
                username=if_credentials.username,
                password=if_credentials.password,
                database=if_credentials.db,
            )
        else:
            logger.error("Influx credentials not enable.")

    def store_data(
        self,
        msg_item: str,
        msg_source: str,
        msg_body: dict,
        msg_time: datetime,
        msg_crc_status: bool,
    ) -> None:
        if not msg_crc_status:
            # CRC failed, message was corrupted. Do not add to database
            logger.debug(
                "CRC FAILED for %s/%s at %s",
                msg_source,
                msg_item,
                msg_time.strftime("%Y-%m-%d %H:%M:%S"),
            )
            return

        # Write data and check if successful
        body = [
            {
                # NOTE: Should check format.
                # Old format was "measurement": msgSource but assumes that
                # all fields in all items are uniquely named
                "measurement": msg_source + "/" + msg_item,
                "time": msg_time.strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
                # dictionary of all fields and corresponding values in CAN message
                "fields": msg_body,
            }
        ]
        influx_success = self.influx_client.write_points(
            body, time_precision="ms", protocol="json"
        )
        if influx_success is False:
            logger.warning(
                "Error writing to Influx for %s/%s (%s)", msg_source, msg_item, msg_body
            )

    def close(self) -> None:
        logger.info("Closing InfluxDB connection")
        self.influx_client.close()
