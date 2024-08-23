import logging

from datetime import datetime
from openpyxl import load_workbook, Workbook
from os.path import exists as fileExists

from influxdb import InfluxDBClient
from Receiver.telemetry_parser3 import TelemetryParser
from Receiver.receiver_config import ifCredentials, xlsxOutputFile

logger = logging.getLogger(__name__)
logger.addHandler(logging.NullHandler())


class TelemetryStorer:
    def __init__(self, parser, storage_plugin_list=None):
        self.parser = parser
        self.storage_plugin_list = storage_plugin_list

    def storeData(self, msg: bytearray) -> None:
        msgItem, msgSource, msgBody, msgTime, msgCRCStatus = self.parser.translate_msg(
            msg)  # this implicitly updates timestamp. I.e always run this first
        for storage_plugin in self.storage_plugin_list:
            storage_plugin.storeData(msgItem, msgSource, msgBody, msgTime,
                                     msgCRCStatus)

    def endSession(self):
        for storage_class in self.storage_plugin_list:
            storage_class.close()


class ExcelStorer:
    def __init__(self, xlsxOutputFile: str):
        self.xlsxOutputFile = xlsxOutputFile
        self.XlsxOutRowPointer: int = 2  # skip first row as that is for column labels
        if not (self.xlsxOutputFile == ''):  # if '', disable xlsx output
            if fileExists(self.xlsxOutputFile):
                self.XlsxOutWorkbook = load_workbook(self.xlsxOutputFile,
                                                     read_only=False)
                if 'Translated Messages' in self.XlsxOutWorkbook.sheetnames:
                    self.XlsxOutWorkSheet = self.XlsxOutWorkbook[
                        'Translated Messages']
                    self.XlsxOutRowPointer = self.XlsxOutWorkSheet.max_row + 3  # leave gap in between sessions. Just to make reader aware that more than one session continued with the same file
            else:
                self.XlsxOutWorkbook = Workbook()  # creates new workbook
                self.XlsxOutWorkSheet = self.XlsxOutWorkbook.create_sheet(
                    title='Translated Messages')

            # If new log, add columns labels and excel filters
            if self.XlsxOutRowPointer == 2:  # if true, then this is a new log
                for i, label in enumerate(['Date', 'Time', 'Source', 'Item',
                                           'Data...']):  # Column labels
                    self.XlsxOutWorkSheet.cell(column=i + 1, row=1, value=label)
                self.XlsxOutWorkSheet.cell(column=21, row=1,
                                           value="CRC check")  # longest record is 19 cells so put CRC in 20th
                self.XlsxOutWorkSheet.auto_filter.ref = 'A1:T' + str(
                    self.XlsxOutWorkSheet.max_row)  # Add filters to columns (Source and Item)
        else:
            logger.warning("No excel output path provided.")

    def storeData(self, msgItem: str, msgSource: str, msgBody: dict,
                  msgTime: datetime,
                  msgCRCStatus: bool) -> None:
        XlsxOutWorkbook = self.XlsxOutWorkbook
        XlsxOutWorkSheet = self.XlsxOutWorkSheet
        XlsxOutRowPointer = self.XlsxOutRowPointer
        xlsxOutputFile = self.xlsxOutputFile

        msgTime = msgTime.replace(tzinfo=None)
        XlsxOutWorkSheet.cell(column=1, row=XlsxOutRowPointer, value=msgTime)
        XlsxOutWorkSheet.cell(column=2, row=XlsxOutRowPointer, value=msgTime)

        XlsxOutWorkSheet.cell(column=3, row=XlsxOutRowPointer, value=msgSource)
        XlsxOutWorkSheet.cell(column=4, row=XlsxOutRowPointer, value=msgItem)

        columnPointer: int = 5
        for dataLabel, value in msgBody.items():
            XlsxOutWorkSheet.cell(column=columnPointer, row=XlsxOutRowPointer,
                                  value=dataLabel)
            # print("DEBUG: " + dataLabel + " : " + str(value))
            XlsxOutWorkSheet.cell(column=columnPointer + 1,
                                  row=XlsxOutRowPointer, value=value)
            columnPointer += 2

        XlsxOutWorkSheet.cell(column=21, row=XlsxOutRowPointer,
                              value=msgCRCStatus)

        self.XlsxOutRowPointer = XlsxOutRowPointer + 1
        XlsxOutWorkbook.save(
            xlsxOutputFile)  # save every time a line is written. NOTE: this can corrupt xlsx file if program doesn't exit gracefully (i.e. the computer is unplugged). This is fine as it can be rebuilt using telemetry hex dump on SD card

    def close(self):
        print("Closing xlsx output file")
        self.XlsxOutWorkbook.save(self.xlsxOutputFile)
        self.XlsxOutWorkbook.close()


class Influx2Storer:
    def __init__(self, ifCredentials):
        if ifCredentials.enabled:  # if false, do not initialise or enable influx output
            self.influxClient = InfluxDBClient(host=ifCredentials.host,
                                               port=ifCredentials.port,
                                               username=ifCredentials.username,
                                               password=ifCredentials.password,
                                               database=ifCredentials.db)
        else:
            logger.warning("Influx credentials not enable.")

    def storeData(self, msgItem: str, msgSource: str, msgBody: dict,
                  msgTime: datetime,
                  msgCRCStatus: bool) -> None:
        if not msgCRCStatus:  # CRC failed, message was corrupted. Do not add to database
            # print("CRC FAILED for " + msgSource + "/" + msgItem + " at " + msgTime.strftime("%Y-%m-%d %H:%M:%S"))
            return

        body = [{
            "measurement": msgSource + '/' + msgItem,
            # NOTE: Should check format. Old format was "measurement": msgSource but assumes that all fields in all items are uniquely named
            # "time": int(msgTime.timestamp() * 1000), #NOTE: get timestamp from 1970 in milliseconds
            "time": msgTime.strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
            "fields": msgBody
            # dictionary of all fields and corresponding values in CAN message
        }]
        influx_success = self.influxClient.write_points(body,
                                                        time_precision='ms',
                                                        protocol='json')  # Write data and check if successful
        if influx_success is False:
            print(
                "Error writing to Influx for %s/%s (%s)" % msgSource % msgItem % msgBody)

    def close(self):
        print("Closing InfluxDB connection")
        self.influxClient.close()


telemetry_parser = TelemetryParser()


class StorerWrapper:
    def __init__(self):
        self.is_storer_init = False
        self.telemetry_storer = None

    def init_storer(self):
        storeFunctionList = []
        if not xlsxOutputFile == '':
            excel_storer = ExcelStorer(xlsxOutputFile)
            storeFunctionList.append(excel_storer)
        if ifCredentials.enabled:
            influx_storer = Influx2Storer(ifCredentials)
            storeFunctionList.append(influx_storer)
        self.telemetry_storer = TelemetryStorer(telemetry_parser,
                                                storage_plugin_list=storeFunctionList)

    def storeData(self, *args):
        if not self.is_storer_init:
            self.init_storer()
        self.telemetry_storer.storeData(*args)

    def endSession(self):
        if not self.is_storer_init:
            self.init_storer()
        self.telemetry_storer.endSession()


storer_wrapper = StorerWrapper()


def storeData(*args):
    storer_wrapper.storeData(*args)


def endSession():
    storer_wrapper.endSession()
